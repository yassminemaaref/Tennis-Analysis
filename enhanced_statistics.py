import numpy as np
import cv2
import pandas as pd
from collections import defaultdict
import json

class EnhancedTennisStatistics:
    """
    Enhanced statistics tracker integrated with existing tennis analysis project.
    Tracks court positioning, shot types, serves, and advanced metrics including rally-by-rally analysis.
    """
    
    def __init__(self, court_keypoints):
        """
        Initialize the statistics tracker.
        
        Args:
            court_keypoints: Array of 28 values [x1,y1,x2,y2,...] from CourtLineDetector
        """
        self.court_keypoints = court_keypoints
        self.court_center_x = None
        self.court_top_y = None
        self.court_bottom_y = None
        self.baseline_zones = {}
        
        self._calculate_court_references()
        
        # Enhanced player statistics
        self.player_stats = {
            1: self._init_player_stats(),
            2: self._init_player_stats()
        }
        
        # Rally tracking
        self.rallies = []
        self.current_rally = []
        self.rally_start_frame = None
        self.rally_serving_player = None
        self.rally_player_positions = {1: [], 2: []}
        self.last_ball_hitter = None
        
        # Frame-by-frame tracking
        self.frame_stats = []
        
    def _init_player_stats(self):
        """Initialize comprehensive statistics for a player."""
        return {
            # Shot statistics
            'total_shots': 0,
            'forehand_shots': 0,
            'backhand_shots': 0,
            'estimated_forehand': 0,
            'estimated_backhand': 0,
            'serves': 0,
            
            # Court positioning
            'left_court_frames': 0,
            'right_court_frames': 0,
            'front_court_frames': 0,
            'back_court_frames': 0,
            
            # Position tracking
            'positions': [],
            'shot_locations': [],
            
            # Movement
            'total_distance_pixels': 0.0,
            'total_distance_meters': 0.0,
            'speeds': [],
            
            # Rally statistics
            'rallies_won': 0,
            'rallies_lost': 0,
            'longest_rally_won': 0,
            
            # Shot patterns
            'crosscourt_shots': 0,
            'down_the_line_shots': 0,
        }
    
    def _calculate_court_references(self):
        """Calculate reference points for court analysis from 14 keypoints."""
        if len(self.court_keypoints) >= 28:
            keypoints_2d = []
            for i in range(0, len(self.court_keypoints), 2):
                keypoints_2d.append((self.court_keypoints[i], self.court_keypoints[i+1]))
            
            all_x = [p[0] for p in keypoints_2d]
            all_y = [p[1] for p in keypoints_2d]
            
            self.court_center_x = np.mean(all_x)
            self.court_top_y = min(all_y)
            self.court_bottom_y = max(all_y)
            
            court_height = self.court_bottom_y - self.court_top_y
            self.baseline_threshold = self.court_bottom_y - (court_height * 0.25)
            self.net_threshold = self.court_top_y + (court_height * 0.25)
    
    def update_frame_stats(self, frame_num, player_detections, ball_detections, 
                          player_mini_court_positions, ball_mini_court_position):
        """Update statistics for current frame."""
        frame_data = {
            'frame_num': frame_num,
            'player_1_pos': None,
            'player_2_pos': None,
            'ball_pos': ball_mini_court_position
        }
        
        for player_id in [1, 2]:
            if player_id in player_detections and player_id in player_mini_court_positions:
                bbox = player_detections[player_id]
                mini_pos = player_mini_court_positions[player_id]
                
                player_center_x = (bbox[0] + bbox[2]) / 2
                player_center_y = (bbox[1] + bbox[3]) / 2
                
                frame_data[f'player_{player_id}_pos'] = mini_pos
                
                self._update_player_position(player_id, player_center_x, player_center_y, 
                                            mini_pos, frame_num)
        
        self.frame_stats.append(frame_data)
    
    def _update_player_position(self, player_id, center_x, center_y, mini_court_pos, frame_num):
        """Update player position and calculate positioning statistics."""
        stats = self.player_stats[player_id]
        
        stats['positions'].append({
            'frame': frame_num,
            'x': center_x,
            'y': center_y,
            'mini_x': mini_court_pos[0],
            'mini_y': mini_court_pos[1]
        })
        
        if self.court_center_x is not None:
            if center_x < self.court_center_x:
                stats['left_court_frames'] += 1
            else:
                stats['right_court_frames'] += 1
        
        if self.court_top_y is not None and self.court_bottom_y is not None:
            if center_y < self.net_threshold:
                stats['front_court_frames'] += 1
            elif center_y > self.baseline_threshold:
                stats['back_court_frames'] += 1
        
        if len(stats['positions']) > 1:
            prev = stats['positions'][-2]
            distance = np.sqrt((center_x - prev['x'])**2 + (center_y - prev['y'])**2)
            stats['total_distance_pixels'] += distance
    
    def start_new_rally(self, serving_player):
        """Start a new rally."""
        self.current_rally = []
        self.rally_serving_player = serving_player
        self.rally_start_frame = None
        self.rally_player_positions = {1: [], 2: []}
    
    def add_shot_to_rally(self, frame_num, player_id, player_position, ball_position, shot_speed):
        """Add a shot to the current rally."""
        if self.rally_start_frame is None:
            self.rally_start_frame = frame_num
        
        self.current_rally.append({
            'frame': frame_num,
            'player': player_id,
            'player_position': player_position,
            'ball_position': ball_position,
            'shot_speed': shot_speed,
            'shot_number': len(self.current_rally) + 1
        })
        
        self.rally_player_positions[player_id].append({
            'frame': frame_num,
            'position': player_position
        })
    
    def end_rally(self, winner_id=None, rally_end_frame=None):
        """Mark end of rally and record detailed statistics."""
        if len(self.current_rally) > 0:
            rally_duration_frames = rally_end_frame - self.rally_start_frame if rally_end_frame else 0
            
            player_1_shots = sum(1 for shot in self.current_rally if shot['player'] == 1)
            player_2_shots = sum(1 for shot in self.current_rally if shot['player'] == 2)
            
            player_1_distance = self._calculate_rally_distance(1)
            player_2_distance = self._calculate_rally_distance(2)
            
            shot_speeds = [shot['shot_speed'] for shot in self.current_rally]
            
            rally_data = {
                'rally_number': len(self.rallies) + 1,
                'shots': self.current_rally.copy(),
                'total_shots': len(self.current_rally),
                'winner': winner_id,
                'serving_player': self.rally_serving_player,
                'duration_frames': rally_duration_frames,
                'duration_seconds': rally_duration_frames / 24.0,
                'player_1_shots': player_1_shots,
                'player_2_shots': player_2_shots,
                'player_1_distance': player_1_distance,
                'player_2_distance': player_2_distance,
                'average_shot_speed': float(np.mean(shot_speeds)) if shot_speeds else 0,
                'max_shot_speed': float(max(shot_speeds)) if shot_speeds else 0,
                'min_shot_speed': float(min(shot_speeds)) if shot_speeds else 0,
                'start_frame': self.rally_start_frame,
                'end_frame': rally_end_frame
            }
            
            self.rallies.append(rally_data)
            
            if winner_id and winner_id in self.player_stats:
                self.player_stats[winner_id]['rallies_won'] += 1
                if len(self.current_rally) > self.player_stats[winner_id]['longest_rally_won']:
                    self.player_stats[winner_id]['longest_rally_won'] = len(self.current_rally)
                
                loser_id = 2 if winner_id == 1 else 1
                if loser_id in self.player_stats:
                    self.player_stats[loser_id]['rallies_lost'] += 1
            
            self.current_rally = []
            self.rally_start_frame = None
            self.rally_player_positions = {1: [], 2: []}
            self.rally_serving_player = None
    
    def _calculate_rally_distance(self, player_id):
        """Calculate distance covered by a player during current rally."""
        positions = self.rally_player_positions.get(player_id, [])
        
        if len(positions) < 2:
            return 0.0
        
        total_distance = 0
        for i in range(1, len(positions)):
            prev = positions[i-1]['position']
            curr = positions[i]['position']
            distance = np.sqrt((curr[0] - prev[0])**2 + (curr[1] - prev[1])**2)
            total_distance += distance
        
        return float(total_distance * 0.05)
    
    def analyze_shot(self, frame_num, player_shot_ball, ball_position, 
                    player_position, opponent_position, ball_speed):
        """Analyze a shot and update statistics."""
        if player_shot_ball not in self.player_stats:
            return
        
        stats = self.player_stats[player_shot_ball]
        stats['total_shots'] += 1
        
        stats['shot_locations'].append({
            'frame': frame_num,
            'ball_pos': ball_position,
            'player_pos': player_position,
            'speed': ball_speed
        })
        
        shot_type = self._estimate_shot_type(player_position, ball_position, player_shot_ball)
        
        if shot_type == 'forehand':
            stats['estimated_forehand'] += 1
        elif shot_type == 'backhand':
            stats['estimated_backhand'] += 1
        
        if self._is_serve(player_position):
            stats['serves'] += 1
        
        shot_direction = self._analyze_shot_direction(player_position, ball_position, opponent_position)
        if shot_direction == 'crosscourt':
            stats['crosscourt_shots'] += 1
        elif shot_direction == 'down_the_line':
            stats['down_the_line_shots'] += 1
        
        self.last_ball_hitter = player_shot_ball
    
    def _estimate_shot_type(self, player_pos, ball_pos, player_id):
        """Estimate if shot is forehand or backhand."""
        ball_relative_x = ball_pos[0] - player_pos[0]
        
        if abs(ball_relative_x) < 20:
            return 'unknown'
        
        if ball_relative_x > 0:
            return 'forehand'
        else:
            return 'backhand'
    
    def _is_serve(self, player_position):
        """Detect if shot is a serve based on player position near baseline."""
        return player_position[1] > 350 or player_position[1] < 50
    
    def _analyze_shot_direction(self, player_pos, ball_pos, opponent_pos):
        """Analyze if shot is crosscourt or down the line."""
        player_side = 'left' if player_pos[0] < 200 else 'right'
        ball_target_side = 'left' if ball_pos[0] < 200 else 'right'
        
        if player_side != ball_target_side:
            return 'crosscourt'
        else:
            return 'down_the_line'
    
    def calculate_distances_in_meters(self, mini_court_width_meters=10.97):
        """Convert pixel distances to real-world meters."""
        for player_id, stats in self.player_stats.items():
            if len(stats['positions']) > 1:
                total_distance_mini = 0
                for i in range(1, len(stats['positions'])):
                    prev = stats['positions'][i-1]
                    curr = stats['positions'][i]
                    dist = np.sqrt((curr['mini_x'] - prev['mini_x'])**2 + 
                                 (curr['mini_y'] - prev['mini_y'])**2)
                    total_distance_mini += dist
                
                stats['total_distance_meters'] = total_distance_mini * 0.05
    
    def calculate_speed_stats(self, fps=24):
        """Calculate player speed statistics."""
        for player_id, stats in self.player_stats.items():
            positions = stats['positions']
            
            if len(positions) < 2:
                continue
            
            speeds = []
            for i in range(1, len(positions)):
                prev = positions[i-1]
                curr = positions[i]
                
                distance = np.sqrt((curr['x'] - prev['x'])**2 + (curr['y'] - prev['y'])**2)
                time_diff = (curr['frame'] - prev['frame']) / fps
                
                if time_diff > 0:
                    speed = distance / time_diff
                    speeds.append(speed)
            
            stats['speeds'] = speeds
    
    def get_court_positioning_percentage(self):
        """Get percentage of time each player spent on left vs right court."""
        result = {}
        
        for player_id, stats in self.player_stats.items():
            total_frames = stats['left_court_frames'] + stats['right_court_frames']
            front_back_total = stats['front_court_frames'] + stats['back_court_frames']
            
            if total_frames > 0:
                left_pct = (stats['left_court_frames'] / total_frames) * 100
                right_pct = (stats['right_court_frames'] / total_frames) * 100
            else:
                left_pct = right_pct = 0
            
            if front_back_total > 0:
                front_pct = (stats['front_court_frames'] / front_back_total) * 100
                back_pct = (stats['back_court_frames'] / front_back_total) * 100
            else:
                front_pct = back_pct = 0
            
            result[player_id] = {
                'left_court': round(left_pct, 2),
                'right_court': round(right_pct, 2),
                'front_court': round(front_pct, 2),
                'back_court': round(back_pct, 2)
            }
        
        return result
    
    def get_rally_summary(self):
        """Get detailed rally-by-rally summary."""
        return {
            'total_rallies': len(self.rallies),
            'rallies': self.rallies,
            'average_rally_length': float(np.mean([r['total_shots'] for r in self.rallies])) if self.rallies else 0,
            'longest_rally': int(max([r['total_shots'] for r in self.rallies])) if self.rallies else 0,
            'shortest_rally': int(min([r['total_shots'] for r in self.rallies])) if self.rallies else 0,
            'average_rally_duration': float(np.mean([r['duration_seconds'] for r in self.rallies])) if self.rallies else 0
        }
    
    def get_summary(self):
        """Get comprehensive statistics summary."""
        positioning = self.get_court_positioning_percentage()
        
        summary = {
            'match_statistics': {
                'total_rallies': int(len(self.rallies)),
                'total_frames_analyzed': int(len(self.frame_stats))
            },
            'players': {}
        }
        
        for player_id, stats in self.player_stats.items():
            total_shots = int(stats['total_shots'])
            
            player_summary = {
                'total_shots': total_shots,
                'serves': int(stats['serves']),
                'estimated_forehand': int(stats['estimated_forehand']),
                'estimated_backhand': int(stats['estimated_backhand']),
                'forehand_percentage': float(round((stats['estimated_forehand'] / total_shots * 100) 
                                           if total_shots > 0 else 0, 2)),
                'backhand_percentage': float(round((stats['estimated_backhand'] / total_shots * 100) 
                                           if total_shots > 0 else 0, 2)),
                'court_positioning': positioning.get(player_id, {}),
                'total_distance_meters': float(round(stats['total_distance_meters'], 2)),
                'average_speed_pixels_per_sec': float(round(np.mean(stats['speeds']) 
                                                     if stats['speeds'] else 0, 2)),
                'max_speed_pixels_per_sec': float(round(max(stats['speeds']) 
                                                 if stats['speeds'] else 0, 2)),
                'rallies_won': int(stats['rallies_won']),
                'rallies_lost': int(stats['rallies_lost']),
                'longest_rally_won': int(stats['longest_rally_won']),
                'win_rate': float(round((stats['rallies_won'] / 
                                 (stats['rallies_won'] + stats['rallies_lost']) * 100) 
                                if (stats['rallies_won'] + stats['rallies_lost']) > 0 else 0, 2)),
                'crosscourt_shots': int(stats['crosscourt_shots']),
                'down_the_line_shots': int(stats['down_the_line_shots'])
            }
            
            summary['players'][f'player_{player_id}'] = player_summary
        
        return summary
    
    def export_to_json(self, filename='tennis_statistics.json'):
        """Export statistics to JSON file."""
        summary = self.get_summary()
        
        with open(filename, 'w') as f:
            json.dump(summary, f, indent=4)
        
        print(f"Statistics exported to {filename}")
        return summary
    
    def export_detailed_csv(self, filename='tennis_detailed_stats.csv'):
        """Export detailed frame-by-frame statistics to CSV."""
        df = pd.DataFrame(self.frame_stats)
        df.to_csv(filename, index=False)
        print(f"Detailed statistics exported to {filename}")
    
    def export_to_excel_with_charts(self, filename='tennis_statistics.xlsx'):
        """Export statistics to Excel with charts and visualizations."""
        try:
            from openpyxl import Workbook
            from openpyxl.chart import BarChart, Reference
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            wb.remove(wb.active)
            
            ws_summary = wb.create_sheet("Match Summary")
            summary = self.get_summary()
            
            ws_summary['A1'] = "TENNIS MATCH STATISTICS"
            ws_summary['A1'].font = Font(size=16, bold=True)
            ws_summary.merge_cells('A1:D1')
            
            row = 3
            ws_summary[f'A{row}'] = "Match Overview"
            ws_summary[f'A{row}'].font = Font(size=12, bold=True)
            row += 1
            ws_summary[f'A{row}'] = "Total Rallies:"
            ws_summary[f'B{row}'] = summary['match_statistics']['total_rallies']
            row += 1
            ws_summary[f'A{row}'] = "Frames Analyzed:"
            ws_summary[f'B{row}'] = summary['match_statistics']['total_frames_analyzed']
            
            for player_key, player_data in summary['players'].items():
                row += 2
                player_num = player_key.split('_')[1]
                ws_summary[f'A{row}'] = f"PLAYER {player_num}"
                ws_summary[f'A{row}'].font = Font(size=14, bold=True)
                ws_summary[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                ws_summary.merge_cells(f'A{row}:D{row}')
                
                row += 1
                stats_data = [
                    ("Total Shots", player_data['total_shots']),
                    ("Serves", player_data['serves']),
                    ("Forehand", player_data['estimated_forehand']),
                    ("Backhand", player_data['estimated_backhand']),
                    ("Distance Covered (m)", player_data['total_distance_meters']),
                    ("Rallies Won", player_data['rallies_won']),
                    ("Rallies Lost", player_data['rallies_lost']),
                    ("Win Rate (%)", player_data['win_rate']),
                ]
                
                for stat_name, stat_value in stats_data:
                    ws_summary[f'A{row}'] = stat_name
                    ws_summary[f'B{row}'] = stat_value
                    row += 1
            
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 15
            
            wb.save(filename)
            print(f"Excel file with charts exported to {filename}")
            
        except Exception as e:
            print(f"Error creating Excel file: {e}")
    
    def draw_enhanced_overlay(self, frame, player_id, frame_num):
        """Draw enhanced statistics overlay on video frame with real-time updates."""
        if player_id not in self.player_stats:
            return frame
        
        current_stats = self._get_stats_up_to_frame(player_id, frame_num)
        positioning = self._get_positioning_up_to_frame(player_id, frame_num)
        
        y_start = 30 if player_id == 1 else 330
        x_start = 20
        box_width = 320
        box_height = 280
        
        overlay = frame.copy()
        cv2.rectangle(overlay, (x_start - 10, y_start - 20), 
                     (x_start + box_width, y_start + box_height), 
                     (0, 0, 0), -1)
        cv2.addWeighted(overlay, 0.7, frame, 0.3, 0, frame)
        
        color = (0, 255, 0) if player_id == 1 else (255, 0, 255)
        cv2.rectangle(frame, (x_start - 10, y_start - 20), 
                     (x_start + box_width, y_start + box_height), 
                     color, 3)
        
        cv2.putText(frame, f"PLAYER {player_id} STATS", 
                   (x_start, y_start), 
                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)
        
        y_pos = y_start + 30
        line_height = 23
        
        total_shots = current_stats['total_shots']
        fg_pct = (current_stats['forehand'] / total_shots * 100) if total_shots > 0 else 0
        bh_pct = (current_stats['backhand'] / total_shots * 100) if total_shots > 0 else 0
        
        stats_text = [
            f"Total Shots: {current_stats['total_shots']}",
            f"Serves: {current_stats['serves']}",
            f"Forehand: {current_stats['forehand']} ({fg_pct:.1f}%)",
            f"Backhand: {current_stats['backhand']} ({bh_pct:.1f}%)",
            "",
            f"Court Position:",
            f"  Left: {positioning['left']:.1f}% | Right: {positioning['right']:.1f}%",
            f"  Front: {positioning['front']:.1f}% | Back: {positioning['back']:.1f}%",
            "",
            f"Distance: {current_stats['distance']:.1f}m",
            f"Rallies: W{current_stats['rallies_won']}/L{current_stats['rallies_lost']}",
            f"Win Rate: {current_stats['win_rate']:.1f}%"
        ]
        
        for i, text in enumerate(stats_text):
            if text == "":
                y_pos += line_height // 2
                continue
            cv2.putText(frame, text, (x_start, y_pos), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
            y_pos += line_height
        
        return frame
    
    def _get_stats_up_to_frame(self, player_id, frame_num):
        """Get statistics calculated only up to the specified frame."""
        stats = self.player_stats[player_id]
        
        total_shots = sum(1 for shot in stats['shot_locations'] if shot['frame'] <= frame_num)
        serves = sum(1 for shot in stats['shot_locations'] 
                    if shot['frame'] <= frame_num and self._is_serve_from_shot(shot))
        
        forehand = 0
        backhand = 0
        for shot in stats['shot_locations']:
            if shot['frame'] <= frame_num:
                shot_type = self._estimate_shot_type(
                    shot['player_pos'], 
                    shot['ball_pos'], 
                    player_id
                )
                if shot_type == 'forehand':
                    forehand += 1
                elif shot_type == 'backhand':
                    backhand += 1
        
        positions_up_to_frame = [p for p in stats['positions'] if p['frame'] <= frame_num]
        distance = 0
        for i in range(1, len(positions_up_to_frame)):
            prev = positions_up_to_frame[i-1]
            curr = positions_up_to_frame[i]
            dist = np.sqrt((curr['mini_x'] - prev['mini_x'])**2 + 
                         (curr['mini_y'] - prev['mini_y'])**2)
            distance += dist
        distance_meters = distance * 0.05
        
        rallies_won = stats['rallies_won']
        rallies_lost = stats['rallies_lost']
        win_rate = (rallies_won / (rallies_won + rallies_lost) * 100) if (rallies_won + rallies_lost) > 0 else 0
        
        return {
            'total_shots': total_shots,
            'serves': serves,
            'forehand': forehand,
            'backhand': backhand,
            'distance': distance_meters,
            'rallies_won': rallies_won,
            'rallies_lost': rallies_lost,
            'win_rate': win_rate
        }
    
    def _get_positioning_up_to_frame(self, player_id, frame_num):
        """Get court positioning stats up to the specified frame."""
        stats = self.player_stats[player_id]
        
        positions_up_to_frame = [p for p in stats['positions'] if p['frame'] <= frame_num]
        
        if len(positions_up_to_frame) == 0:
            return {'left': 0, 'right': 0, 'front': 0, 'back': 0}
        
        left_count = 0
        right_count = 0
        front_count = 0
        back_count = 0
        
        for pos in positions_up_to_frame:
            if self.court_center_x is not None:
                if pos['x'] < self.court_center_x:
                    left_count += 1
                else:
                    right_count += 1
            
            if self.court_top_y is not None and self.court_bottom_y is not None:
                if pos['y'] < self.net_threshold:
                    front_count += 1
                elif pos['y'] > self.baseline_threshold:
                    back_count += 1
        
        total_lr = left_count + right_count
        total_fb = front_count + back_count
        
        return {
            'left': (left_count / total_lr * 100) if total_lr > 0 else 0,
            'right': (right_count / total_lr * 100) if total_lr > 0 else 0,
            'front': (front_count / total_fb * 100) if total_fb > 0 else 0,
            'back': (back_count / total_fb * 100) if total_fb > 0 else 0
        }
    
    def _is_serve_from_shot(self, shot):
        """Check if a shot location indicates a serve."""
        return shot['player_pos'][1] > 350 or shot['player_pos'][1] < 50
    
    def print_summary(self):
        """Print formatted summary to console."""
        summary = self.get_summary()
        
        print("\n" + "="*60)
        print("TENNIS MATCH STATISTICS SUMMARY")
        print("="*60)
        
        print(f"\nMatch Overview:")
        print(f"  Total Rallies: {summary['match_statistics']['total_rallies']}")
        print(f"  Frames Analyzed: {summary['match_statistics']['total_frames_analyzed']}")
        
        for player_key, player_data in summary['players'].items():
            player_num = player_key.split('_')[1]
            print(f"\n{'='*60}")
            print(f"PLAYER {player_num}")
            print(f"{'='*60}")
            
            print(f"\n  Shot Statistics:")
            print(f"    Total Shots: {player_data['total_shots']}")
            print(f"    Serves: {player_data['serves']}")
            print(f"    Forehand (est): {player_data['estimated_forehand']} ({player_data['forehand_percentage']}%)")
            print(f"    Backhand (est): {player_data['estimated_backhand']} ({player_data['backhand_percentage']}%)")
            
            print(f"\n  Court Positioning:")
            pos = player_data['court_positioning']
            print(f"    Left Court: {pos['left_court']}% | Right Court: {pos['right_court']}%")
            print(f"    Front Court: {pos['front_court']}% | Back Court: {pos['back_court']}%")
            
            print(f"\n  Movement:")
            print(f"    Distance Covered: {player_data['total_distance_meters']:.2f} meters")
            
            print(f"\n  Rally Performance:")
            print(f"    Wins: {player_data['rallies_won']} | Losses: {player_data['rallies_lost']}")
            print(f"    Win Rate: {player_data['win_rate']}%")
            print(f"    Longest Rally Won: {player_data['longest_rally_won']} shots")
        
        print("\n" + "="*60)